#!/usr/bin/env python3
"""parse-prices.py — полный парсер прайса для всех категорий"""

from openpyxl import load_workbook
from collections import defaultdict, OrderedDict
from pathlib import Path
import re

XLSX = '/mnt/user-data/uploads/прайс_apple_collecty.xlsx'
OUT_DIR = '/home/claude/epl-collecty/src/lib/generated'

# ─── Цвета ────────────────────────────────────────────────────────────────
COLOR_MAP = {
    # iPhone
    'orange':     ('Оранжевый (Cosmic Orange)', '#D55A2B'),
    'blue':       ('Синий',                     '#3B5D78'),
    'silver':     ('Серебристый',               '#D8D8D8'),
    'black':      ('Чёрный',                    '#1C1C1E'),
    'white':      ('Белый',                     '#F2F1ED'),
    'sage':       ('Шалфей',                    '#A8B19E'),
    'lavender':   ('Лавандовый',                '#C9B8D4'),
    'lavander':   ('Лавандовый',                '#C9B8D4'),
    'desert':     ('Пустынный Титан',           '#C8A97E'),
    'natural':    ('Натуральный Титан',         '#B5A99A'),
    'pink':       ('Розовый',                   '#F6D6D4'),
    'teal':       ('Бирюзовый',                 '#A5C8C6'),
    'ultramarine':('Ультрамариновый',           '#5B6FB3'),
    'green':      ('Зелёный',                   '#C5D0BC'),
    'yellow':     ('Жёлтый',                    '#FAEACB'),
    'gray':       ('Серый',                     '#4E4E4F'),
    'graphite':   ('Графитовый',                '#3C3C3E'),
    'lightgray':  ('Светло-серый',              '#B0B2B4'),
    'starlight':  ('Сияющая Звезда',            '#F5EFE3'),
    'skyblue':    ('Небесно-голубой',           '#BDD3E0'),
    'sky blue':   ('Небесно-голубой',           '#BDD3E0'),
    'silverblue': ('Серебристо-голубой',        '#B5C5D8'),
    'iceblue':    ('Ледяной голубой',           '#C8DFE5'),
    'midnight':   ('Тёмная Ночь',               '#2C2D30'),
    'jetblack':   ('Чёрный (Jet Black)',        '#0E0E10'),
    'navy':       ('Тёмно-синий (Navy)',        '#23314D'),
    'mint':       ('Мятный',                    '#B5E0CC'),
    'olive':      ('Оливковый',                 '#7E8265'),
    'coralred':   ('Коралловый красный',        '#E26B5A'),
    'coral red':  ('Коралловый красный',        '#E26B5A'),
    'whitesilver':('Бело-серебристый',          '#E8E8EA'),
    'indigo':     ('Индиго',                    '#4B4D76'),
    'citrus':     ('Цитрусовый',                '#DDD389'),
    'purple':     ('Фиолетовый',                '#B8A8CF'),
    'violet':     ('Фиолетовый',                '#8B7AAF'),
    # Watch + AirPods
    'rose':         ('Розовый (Rose)',           '#E6BFC2'),
    'rose gold':    ('Розовое золото',           '#E0BFB8'),
    'space gray':   ('Серый космос',             '#4E4E4F'),
    'jet black':    ('Глубокий чёрный',          '#0E0E10'),
    'cream':        ('Кремовый',                 '#EFE5D2'),
    'brown':        ('Коричневый',               '#5C3B29'),
    # Galaxy
    'jasper plum':  ('Сливовый',                 '#5A3A4A'),
    'vinca blue':   ('Тёмно-синий',              '#3D5A82'),
    'prussian blue':('Прусский синий',           '#1F3354'),
    'ceramic pink': ('Керамический розовый',     '#E8B5A8'),
    'apricot topaz':('Абрикосовый топаз',        '#D4A88A'),
    'amber silk':   ('Янтарный шёлк',            '#C49862'),
    'red velvet':   ('Красный бархат',           '#86323A'),
    'nickel/cooper':('Никель/медь',              '#A07B5C'),
    'strawberry/bronze': ('Клубника/бронза',     '#B0716E'),
    'yellow/nickel':('Жёлтый/никель',            '#D4C56E'),
    'mat black/cooper': ('Матовый чёрный/медь',  '#2A2A2C'),
    'silver/nickel':('Серебро/никель',           '#C7C7C9'),
    'blue/cooper':  ('Синий/медь',               '#5A6F8A'),
}

SIM_MAP = {
    'eSIM':     ('esim',     'eSIM',             'Только eSIM, без физической SIM. Нужна поддержка eSIM у оператора.'),
    'SIM+eSIM': ('sim-esim', 'Nano-SIM + eSIM',  'Европейская/РФ версия. Одна физическая SIM + одна eSIM.'),
    'wi-fi':    ('wifi',     'Wi-Fi',            'Только Wi-Fi, без сотовой связи.'),
    'LTE':      ('lte',      'Wi-Fi + Cellular', 'Wi-Fi и сотовая связь (LTE/5G) с eSIM.'),
}

DEFAULT_SIM_IPHONE = ('sim-esim', 'Nano-SIM + eSIM', 'Европейская/РФ версия.')
DEFAULT_SIM_WIFI   = ('wifi', 'Wi-Fi', 'Только Wi-Fi, без сотовой связи.')
DEFAULT_SIM_NONE   = ('none', 'Стандарт', 'Стандартная комплектация.')


def ts_str(s):
    if s is None: s = ""
    return '"' + str(s).replace('\\', '\\\\').replace('"', '\\"').replace('\n', '\\n') + '"'


def get_color(name):
    if not name:
        return ('Стандартный', '#888888')
    key = name.strip().lower()
    if key in COLOR_MAP:
        return COLOR_MAP[key]
    # Поиск по подстроке (более длинные ключи приоритетнее)
    for k in sorted(COLOR_MAP.keys(), key=len, reverse=True):
        if k in key:
            return COLOR_MAP[k]
    clean = re.sub(r'[^а-яА-Яa-zA-Z\s]', '', name).strip()
    return (clean.capitalize() or 'Стандартный', '#888888')


def memory_label_gb(gb):
    if gb >= 1024:
        return f"{gb // 1024} ТБ"
    return f"{gb} ГБ"


def parse_memory_gb(s):
    s = s.strip().upper()
    if 'TB' in s:
        return int(s.replace('TB', '').strip()) * 1024
    return int(s.replace('GB', '').strip())


# ─── iPhone ────────────────────────────────────────────────────────────────
def parse_iphone():
    wb = load_workbook(XLSX, data_only=True)
    data = defaultdict(list)

    ws = wb['iPhone 17']
    section = 'iphone-17-pro-max'
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0] or not row[6]: continue
        mem = parse_memory_gb(row[3]); price = row[6]
        if mem == 256 and section == 'iphone-17-pro-max' and price < 108000:
            section = 'iphone-17-pro'
        if mem == 256 and section == 'iphone-17-pro' and price < 90000:
            section = 'iphone-17'
        data[section].append({
            'storage_id': str(mem), 'storage_label': memory_label_gb(mem),
            'color_raw': row[4], 'sim_raw': row[5], 'price': price,
        })

    slug_map_16 = {
        'iPhone 16 Pro Max': 'iphone-16-pro-max', 'iPhone 16 Pro': 'iphone-16-pro',
        'iPhone 16 Plus': 'iphone-16-plus', 'iPhone 16': 'iphone-16', 'iPhone 16E': 'iphone-16e',
    }
    ws = wb['iPhone 16']
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        slug = slug_map_16.get(row[1].strip())
        if not slug: continue
        mem = parse_memory_gb(row[3])
        data[slug].append({
            'storage_id': str(mem), 'storage_label': memory_label_gb(mem),
            'color_raw': row[4], 'sim_raw': row[5], 'price': row[6],
        })

    slug_map_15 = {
        'iPhone 15 Pro Max': 'iphone-15-pro-max', 'iPhone 15 Pro': 'iphone-15-pro',
        'iPhone 15 Plus': 'iphone-15-plus', 'iPhone 15': 'iphone-15',
    }
    ws = wb['iPhone 15']
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        slug = slug_map_15.get(row[1].strip())
        if not slug: continue
        mem = parse_memory_gb(row[3])
        data[slug].append({
            'storage_id': str(mem), 'storage_label': memory_label_gb(mem),
            'color_raw': row[4], 'sim_raw': row[5], 'price': row[6],
        })
    return data


# ─── iPad ──────────────────────────────────────────────────────────────────
def parse_ipad():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['iPad']
    data = defaultdict(list)
    slug_map = {
        'iPad 11 (2025)': 'ipad-11-2025', 'iPad Air 11 M3': 'ipad-air-11-m3',
        'iPad Air 11 M4': 'ipad-air-11-m4', 'iPad Pro 11': 'ipad-pro-11', 'iPad Pro 13': 'ipad-pro-13',
    }
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        slug = slug_map.get(row[1].strip())
        if not slug: continue
        color = row[4] or ''
        for prefix in ['iPad 11 2025', 'iPad Air 11 M3', 'iPad Air 11 M4', 'iPad Pro 11', 'iPad Pro 13', 'M5', 'M4', 'M3']:
            color = color.replace(prefix, '')
        color = color.strip()
        mem = parse_memory_gb(row[3])
        data[slug].append({
            'storage_id': str(mem), 'storage_label': memory_label_gb(mem),
            'color_raw': color, 'sim_raw': row[5], 'price': row[6],
        })
    return data


# ─── MacBook ───────────────────────────────────────────────────────────────
def parse_macbook():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['MacBook']
    data = defaultdict(list)
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        group = row[1].strip()
        variant = (row[4] or '').strip()
        mem_config = (row[3] or '').strip()
        price = row[6]
        chip_suffix = ''
        vlow = variant.lower()
        for c in ['m5 pro', 'm4 pro', 'm5', 'm4', 'm3', 'm2']:
            if c in vlow:
                chip_suffix = '-' + c.replace(' ', '-')
                break
        if group == 'Mac mini':
            slug = f'mac-mini{chip_suffix}' if chip_suffix else 'mac-mini'
        elif group == 'MacBook Neo': slug = 'macbook-neo'
        elif group == 'MacBook Air 13': slug = f'macbook-air-13{chip_suffix}'
        elif group == 'MacBook Air 15': slug = f'macbook-air-15{chip_suffix}'
        elif group == 'MacBook Pro': slug = f'macbook-pro-14{chip_suffix}'
        else: continue
        color = vlow
        for x in ['m5 pro', 'm4 pro', 'm5', 'm4', 'm3', 'm2', '13', '14', '15']:
            color = color.replace(x, '')
        color = color.strip('/ ').strip()
        storage_id = mem_config.replace('/', '-').replace(' ', '')
        if '/' in mem_config:
            ram, ssd = mem_config.split('/')
            ssd_u = ssd.strip().upper()
            ssd_label = ssd_u.replace('TB', ' ТБ') if 'TB' in ssd_u else f"{ssd_u.replace('GB', '').strip()} ГБ"
            storage_label = f"{ram.strip()} ГБ / {ssd_label} SSD"
        else:
            storage_label = mem_config
        data[slug].append({
            'storage_id': storage_id, 'storage_label': storage_label,
            'color_raw': color, 'sim_raw': None, 'price': price,
        })
    return data


# ─── Watch ─────────────────────────────────────────────────────────────────
def parse_watch():
    """Apple Watch и Galaxy Watch: вариант — это полная строка типа 'Watch Ultra 3 black ti black ocean band'"""
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Watch']
    data = defaultdict(list)

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        group = row[1].strip()
        variant = (row[4] or '').strip()
        sim = row[5]
        price = row[6]

        # Определяем slug по содержимому variant
        vlow = variant.lower()
        if group == 'Apple Watch Ultra':
            if 'ultra 3' in vlow: slug = 'apple-watch-ultra-3'
            elif 'ultra 2' in vlow: slug = 'apple-watch-ultra-2'
            elif 'ultra 2025' in vlow or '2025' in vlow: slug = 'apple-watch-ultra-2025'
            else: slug = 'apple-watch-ultra'
        elif group == 'Apple Watch Series 11': slug = 'apple-watch-series-11'
        elif group == 'Apple Watch Series 10': slug = 'apple-watch-series-10'
        elif group == 'Apple Watch SE':
            if 'se 3' in vlow or 'se3' in vlow: slug = 'apple-watch-se-3'
            elif 'se 2' in vlow or 'se2' in vlow: slug = 'apple-watch-se-2'
            else: slug = 'apple-watch-se'
        elif group == 'Samsung Galaxy Watch':
            if 'galaxy watch 8' in vlow or 'samsung galaxy watch 8' in vlow or vlow.startswith('8'): 
                slug = 'samsung-galaxy-watch-8'
            else: slug = 'samsung-galaxy-watch'
        elif group == 'Другое':
            if 'galaxy watch classic 8' in vlow: slug = 'samsung-galaxy-watch-classic-8'
            else: continue
        else: continue

        # Storage не относится к часам — используем размер корпуса как "storage"
        size = '46mm' if '46' in variant else '44mm' if '44' in variant else '42mm' if '42' in variant else '40mm' if '40' in variant else 'std'
        size_label = '46 мм' if '46mm' == size else '44 мм' if '44mm' == size else '42 мм' if '42mm' == size else '40 мм' if '40mm' == size else 'Стандарт'

        # Цвет — берём ключевое слово из variant
        # Удаляем номера моделей и размеры
        color_clean = vlow
        for x in ['watch ultra 3', 'watch ultra 2', 'watch ultra 2025', 'watch series 11', 'watch series 10',
                  'watch se 3 2025', 'watch se 3', 'watch se 2', 'galaxy watch classic 8',
                  'samsung galaxy watch 8', 'galaxy watch 8', '46mm', '44mm', '42mm', '40mm',
                  '46', '44', '42', '40', 'ti', 'mm', '2025']:
            color_clean = color_clean.replace(x, '')
        color_clean = re.sub(r'\s+', ' ', color_clean).strip()
        # Берём первое цветное слово
        color_words = color_clean.split()
        color_raw = color_words[0] if color_words else 'standard'

        data[slug].append({
            'storage_id': size, 'storage_label': size_label,
            'color_raw': color_raw, 'sim_raw': sim, 'price': price,
            'variant_full': variant,
        })
    return data


# ─── AirPods / Наушники ────────────────────────────────────────────────────
def parse_airpods():
    """Все наушники в одну категорию airpods (хотя некоторые Galaxy/Marshall)"""
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Наушники']
    data = defaultdict(list)
    
    slug_map = {
        'AirPods Max': 'airpods-max',
        'AirPods Pro': 'airpods-pro',  # 2 модели: Pro 3 и Pro 2
        'AirPods 4':   'airpods-4',
        'Galaxy Buds': 'galaxy-buds',
        'Marshall':    'marshall-headphones',
    }
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        group = row[1].strip()
        slug = slug_map.get(group)
        if not slug: continue
        variant = (row[4] or '').strip()
        vlow = variant.lower()
        
        # Для AirPods Pro — две модели в одной группе
        if group == 'AirPods Pro':
            if vlow.startswith('3') or 'pro 3' in vlow: slug = 'airpods-pro-3'
            elif '2 usb-c' in vlow or 'pro 2' in vlow: slug = 'airpods-pro-2'
        
        # Цвет
        color_clean = vlow
        for x in ['2024', 'usb-c', 'anc', '5', '4', '3 pro', '3 fe', '3', '2', 'pro', 'silver fe', 'fe']:
            color_clean = color_clean.replace(x, '')
        color_clean = color_clean.strip()
        if not color_clean:
            color_clean = 'standard'
        
        data[slug].append({
            'storage_id': 'std', 'storage_label': 'Стандарт',
            'color_raw': color_clean, 'sim_raw': None, 'price': row[6],
            'variant_full': variant,
        })
    return data


# ─── Android ───────────────────────────────────────────────────────────────
def parse_android():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Android']
    data = defaultdict(list)
    
    slug_map = {
        'Samsung Galaxy S26 Ultra': 'samsung-galaxy-s26-ultra',
        'Samsung Galaxy S26+':      'samsung-galaxy-s26-plus',
        'Samsung Galaxy S26':       'samsung-galaxy-s26',
        'Samsung Galaxy S25 Ultra': 'samsung-galaxy-s25-ultra',
        'Samsung Galaxy S25 FE':    'samsung-galaxy-s25-fe',
        'Samsung Galaxy S25':       'samsung-galaxy-s25',
        'Samsung Galaxy A56':       'samsung-galaxy-a56',
        'Samsung Galaxy A36':       'samsung-galaxy-a36',
        'Samsung Galaxy A26':       'samsung-galaxy-a26',
        'Samsung Galaxy A17':       'samsung-galaxy-a17',
    }
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        group = row[1].strip()
        variant = (row[4] or '').strip()
        mem_config = (row[3] or '').strip()
        price = row[6]
        
        slug = slug_map.get(group)
        
        # Группа "Другое" — определяем по variant
        if group == 'Другое':
            vlow = variant.lower()
            if 'mi 15t' in vlow: slug = 'xiaomi-mi-15t'
            elif 'redmi note 15 pro' in vlow: slug = 'xiaomi-redmi-note-15-pro'
            elif 'redmi note 15' in vlow: slug = 'xiaomi-redmi-note-15'
            elif 'redmi note 14s' in vlow: slug = 'xiaomi-redmi-note-14s'
            elif 'redmi note 14' in vlow: slug = 'xiaomi-redmi-note-14'
            elif 'meizu note 21' in vlow: slug = 'meizu-note-21'
            elif 'galaxy a07' in vlow: slug = 'samsung-galaxy-a07'
            else: continue
            
            # Чистим variant от названия модели
            color_raw = variant
            for prefix in ['Samsung Galaxy A07', 'Mi 15T', 'Xiaomi Redmi Note 15 Pro', 
                          'Xiaomi Redmi Note 15', 'Xiaomi Redmi Note 14S', 'Xiaomi Redmi Note 14',
                          'Meizu Note 21']:
                color_raw = color_raw.replace(prefix, '')
            color_raw = color_raw.strip()
            if not color_raw: color_raw = 'standard'
        else:
            color_raw = variant
        
        if not slug: continue
        
        # Парсим память: "12/512" → storage_id "12-512", label "12 ГБ / 512 ГБ"
        if '/' in mem_config:
            ram, store = mem_config.split('/')
            storage_id = mem_config.replace('/', '-').replace(' ', '')
            store_u = store.strip().upper()
            store_label = store_u.replace('TB', ' ТБ') if 'TB' in store_u else f"{store_u.replace('GB', '').strip()} ГБ"
            storage_label = f"{ram.strip()} ГБ / {store_label}"
        else:
            storage_id = mem_config or 'std'
            storage_label = mem_config or 'Стандарт'
        
        data[slug].append({
            'storage_id': storage_id, 'storage_label': storage_label,
            'color_raw': color_raw.lower(), 'sim_raw': None, 'price': price,
        })
    return data


# ─── Dyson ─────────────────────────────────────────────────────────────────
def parse_dyson():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Dyson']
    data = defaultdict(list)
    
    # Каждый продукт уникален — группируем по типу "Фены/Стайлеры/Выпрямители/Пылесосы"
    # И внутри по модели (HD17/HD18/HS09/HS08/V15 и т.д.)
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        group = row[1].strip()
        variant = (row[4] or '').strip()
        price = row[6]
        vlow = variant.lower()
        
        # Определяем slug по типу + модели
        if group == 'Фены Dyson':
            if 'hd18' in vlow: slug = 'dyson-hd18-pro'
            elif 'hd17' in vlow: slug = 'dyson-hd17'
            elif 'hd16' in vlow: slug = 'dyson-hd16'
            else: slug = 'dyson-fen'
        elif group == 'Стайлеры Dyson':
            if 'hs09' in vlow: slug = 'dyson-airwrap-hs09'
            elif 'hs08' in vlow: slug = 'dyson-airwrap-hs08'
            elif 'hs05' in vlow: slug = 'dyson-airwrap-hs05'
            else: slug = 'dyson-stayler'
        elif group == 'Выпрямители Dyson':
            if 'ht01' in vlow: slug = 'dyson-airstrait-ht01'
            else: slug = 'dyson-vypryamitel'
        elif group == 'Пылесосы Dyson':
            if 'wash g1' in vlow: slug = 'dyson-wash-g1'
            elif 'gen 5' in vlow: slug = 'dyson-gen5-detect'
            elif 'v16s' in vlow: slug = 'dyson-v16s'
            elif 'v15s' in vlow: slug = 'dyson-v15s'
            elif 'v15' in vlow: slug = 'dyson-v15'
            elif 'v12s' in vlow: slug = 'dyson-v12s'
            elif 'v12' in vlow: slug = 'dyson-v12'
            elif 'v8' in vlow: slug = 'dyson-v8'
            else: slug = 'dyson-pylesos'
        else:
            continue
        
        # Цвет — это всё что после кода модели
        color_clean = vlow
        for code in ['hd18 pro', 'hd17', 'hd16', 'hs09 coanda 2x', 'hs08 long id.', 'hs08 long id',
                     'hs05 long.', 'hs05 long', 'ht01', 'wash g1.', 'wash g1',
                     'gen 5 detect absolute', 'v16s piston animal submarine sv53a',
                     'v15s detect submarine sv47', 'v15 detect absolute sv47',
                     'v12 detect slim absolute sv46', 'v12s detect slim submarine',
                     'v8 advanced', '(распак)', '(с кейсом)', '(без кейса)', 'case']:
            color_clean = color_clean.replace(code, '')
        color_clean = re.sub(r'\s+', ' ', color_clean).strip().strip('.').strip()
        if not color_clean: color_clean = 'standard'
        
        data[slug].append({
            'storage_id': 'std', 'storage_label': 'Стандарт',
            'color_raw': color_clean, 'sim_raw': None, 'price': price,
            'variant_full': variant,
        })
    return data


# ─── Акустика и гаджеты ────────────────────────────────────────────────────
def parse_audio():
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Акустика и гаджеты']
    data = defaultdict(list)
    
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[1] or not row[6]: continue
        group = row[1].strip()
        variant = (row[4] or '').strip()
        price = row[6]
        vlow = variant.lower()
        
        if group == 'JBL':
            if 'boombox 4' in vlow: slug = 'jbl-boombox-4'
            else: slug = 'jbl-speaker'
            color = 'standard'
        elif group == 'Яндекс станция':
            slug = 'yandex-station-light-2'
            # цвет: Лайт 2 розовая → розовая
            color = vlow.replace('лайт 2', '').strip()
        elif group == 'DJI Микрофоны':
            if 'mic 3' in vlow: slug = 'dji-mic-3'
            elif 'mic 2' in vlow: slug = 'dji-mic-2'
            elif 'mic mini' in vlow: slug = 'dji-mic-mini'
            else: continue
            color = 'standard'
        else:
            continue
        
        data[slug].append({
            'storage_id': 'std', 'storage_label': 'Стандарт',
            'color_raw': color, 'sim_raw': None, 'price': price,
            'variant_full': variant,
        })
    return data


# ═══════════════════════════════════════════════════════════════════════════
#  ОБЩАЯ build_config — без изменений
# ═══════════════════════════════════════════════════════════════════════════

def build_config(slug, records, category, default_sim):
    colors_seen = OrderedDict()
    for r in records:
        key = (r['color_raw'] or '').strip().lower()
        if key and key not in colors_seen:
            ru, hex_c = get_color(r['color_raw'])
            colors_seen[key] = {
                'id': re.sub(r'[^a-z0-9]', '-', key).strip('-') or 'default',
                'name': ru, 'hex': hex_c, 'image': slug,
            }
    if not colors_seen:
        colors_seen['default'] = {'id': 'default', 'name': 'Стандартный', 'hex': '#888888', 'image': slug}
    colors = list(colors_seen.values())

    storage_seen = OrderedDict()
    for r in records:
        if r['storage_id'] and r['storage_id'] not in storage_seen:
            storage_seen[r['storage_id']] = {'id': r['storage_id'], 'label': r['storage_label'], 'available': True}
    storage = list(storage_seen.values())

    sim_seen = OrderedDict()
    for r in records:
        if r['sim_raw'] and r['sim_raw'] in SIM_MAP:
            sid, lbl, desc = SIM_MAP[r['sim_raw']]
            if sid not in sim_seen:
                sim_seen[sid] = {'id': sid, 'label': lbl, 'description': desc}
    if not sim_seen:
        sid, lbl, desc = default_sim
        sim_seen[sid] = {'id': sid, 'label': lbl, 'description': desc}
    sims = list(sim_seen.values())

    price_map = {}
    for r in records:
        cid = re.sub(r'[^a-z0-9]', '-', (r['color_raw'] or '').strip().lower()).strip('-') or 'default'
        if r['sim_raw'] and r['sim_raw'] in SIM_MAP:
            simid = SIM_MAP[r['sim_raw']][0]
        else:
            simid = default_sim[0]
        k = (r['storage_id'], cid, simid)
        if k not in price_map or r['price'] < price_map[k]:
            price_map[k] = r['price']

    prices = [{'storageId': s, 'colorId': c, 'simId': si, 'price': p}
              for (s, c, si), p in sorted(price_map.items())]

    return {
        'slug': slug, 'category': category,
        'colors': colors, 'storage': storage, 'sim': sims, 'prices': prices,
        'defaultStorage': storage[0]['id'] if storage else '',
        'defaultColor': colors[0]['id'] if colors else '',
        'defaultSim': sims[0]['id'] if sims else '',
        'priceFrom': min(p['price'] for p in prices) if prices else 0,
    }


# ═══════════════════════════════════════════════════════════════════════════
#  МЕТАДАННЫЕ
# ═══════════════════════════════════════════════════════════════════════════

# Спецификации только для топ-моделей. Для остальных — пустые (специально, чтобы блок не показывался)
MODEL_META = {
    # ═══ iPhone ═══════════════════════════════════════════════════════════════
    'iphone-17-pro-max': {
        'compareTitle': 'iPhone 16 Pro Max',
        'specs': [
            ('Процессор', 'Apple A19 Pro (3 нм)'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,9" Super Retina XDR OLED, 120 Гц ProMotion, 3000 нит'),
            ('Основная камера', '48 МП Fusion + 48 МП Ultra Wide + 48 МП Telephoto (8× оптический tetraprism)'),
            ('Фронтальная камера', '18 МП Center Stage'),
            ('Видео', '4K Dolby Vision до 120 fps, ProRes RAW, Apple Log 2'),
            ('Батарея', 'до 33 часов видео'),
            ('Корпус', 'Алюминий Unibody + Ceramic Shield 2'),
            ('Разъём', 'USB-C (USB 3, до 10 Гбит/с)'),
            ('Защита', 'IP68 (6 м, 30 мин)'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
            ('Action Button', 'Да'),
            ('MagSafe', 'До 25 Вт'),
            ('5G', 'Да'),
            ('Габариты', '163,0 × 77,6 × 8,75 мм'),
            ('Вес', '233 г'),
        ],
        'compare': [
            ('Процессор', 'A19 Pro (3 нм)', 'A18 Pro (3 нм)', True),
            ('Оперативная память', '12 ГБ', '8 ГБ', True),
            ('Основная камера', '48+48+48 МП', '48+48+12 МП', True),
            ('Оптический зум', '8× (tetraprism)', '5× (tetraprism)', True),
            ('Фронтальная камера', '18 МП Center Stage', '12 МП', True),
            ('Батарея', 'до 33 ч видео', 'до 33 ч видео', False),
            ('Корпус', 'Алюминий + Ceramic Shield 2', 'Титан + Ceramic Shield', False),
            ('Стекло', 'Ceramic Shield 2', 'Ceramic Shield', True),
            ('Apple Intelligence', 'Да', 'Да', False),
            ('Camera Control', 'Да', 'Да', False),
            ('MagSafe', 'до 25 Вт', 'до 25 Вт', False),
            ('USB-C', 'USB 3 (10 Гбит/с)', 'USB 3 (10 Гбит/с)', False),
            ('Цена от', 'от 108 900 ₽', 'от 94 000 ₽', False),
        ],
    },
    'iphone-17-pro': {
        'compareTitle': 'iPhone 16 Pro',
        'specs': [
            ('Процессор', 'Apple A19 Pro (3 нм)'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,3" Super Retina XDR OLED, 120 Гц ProMotion, 3000 нит'),
            ('Основная камера', '48 МП Fusion + 48 МП Ultra Wide + 48 МП Telephoto (8× оптический tetraprism)'),
            ('Фронтальная камера', '18 МП Center Stage'),
            ('Видео', '4K Dolby Vision до 120 fps, ProRes, Apple Log 2'),
            ('Батарея', 'до 31 часов видео'),
            ('Корпус', 'Алюминий Unibody + Ceramic Shield 2'),
            ('Разъём', 'USB-C (USB 3, до 10 Гбит/с)'),
            ('Защита', 'IP68'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
            ('Action Button', 'Да'),
            ('MagSafe', 'До 25 Вт'),
        ],
        'compare': [
            ('Процессор', 'A19 Pro', 'A18 Pro', True),
            ('Оперативная память', '12 ГБ', '8 ГБ', True),
            ('Основная камера', '48+48+48 МП', '48+48+12 МП', True),
            ('Фронтальная камера', '18 МП', '12 МП', True),
            ('Оптический зум', '8× (tetraprism)', '5× (tetraprism)', True),
            ('Батарея', 'до 31 ч видео', 'до 27 ч видео', True),
            ('Стекло', 'Ceramic Shield 2', 'Ceramic Shield', True),
            ('Цена от', 'от 101 300 ₽', 'от 80 700 ₽', False),
        ],
    },
    'iphone-17': {
        'compareTitle': 'iPhone 16',
        'specs': [
            ('Процессор', 'Apple A19'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '6,3" Super Retina XDR OLED, 120 Гц ProMotion'),
            ('Основная камера', '48 МП Fusion + 12 МП Ultra Wide'),
            ('Фронтальная камера', '18 МП Center Stage'),
            ('Батарея', 'до 30 часов видео'),
            ('Корпус', 'Алюминий'),
            ('Разъём', 'USB-C'),
            ('Защита', 'IP68'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
        ],
        'compare': [
            ('Процессор', 'A19', 'A18', True),
            ('Дисплей', '6,3" 120 Гц ProMotion', '6,1" 60 Гц', True),
            ('Всегда включённый дисплей', 'Да', 'Нет', True),
            ('Фронтальная камера', '18 МП', '12 МП', True),
            ('Батарея', 'до 30 ч видео', 'до 22 ч видео', True),
            ('Оперативная память', '8 ГБ', '8 ГБ', False),
            ('Цена от', 'от 66 000 ₽', 'от 57 500 ₽', False),
        ],
    },
    'iphone-16-pro-max': {
        'compareTitle': 'iPhone 15 Pro Max',
        'specs': [
            ('Процессор', 'Apple A18 Pro (3 нм, 2-е поколение)'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '6,9" Super Retina XDR OLED, 120 Гц ProMotion'),
            ('Основная камера', '48 МП + 48 МП Ultra Wide + 12 МП Telephoto (5× зум)'),
            ('Фронтальная камера', '12 МП TrueDepth'),
            ('Батарея', 'до 33 часов видео'),
            ('Корпус', 'Титан (5 поколение)'),
            ('Разъём', 'USB-C (USB 3, до 10 Гбит/с)'),
            ('Защита', 'IP68 (6 м, 30 мин)'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
            ('MagSafe', 'До 25 Вт'),
        ],
        'compare': [
            ('Процессор', 'A18 Pro (3 нм 2-е пок.)', 'A17 Pro (3 нм)', True),
            ('Оперативная память', '8 ГБ', '8 ГБ', False),
            ('Диагональ экрана', '6,9"', '6,7"', True),
            ('Оптический зум', '5×', '5×', False),
            ('Ultra Wide', '48 МП', '12 МП', True),
            ('Camera Control', 'Да', 'Нет', True),
            ('Apple Intelligence', 'Да', 'Ограниченная', True),
            ('MagSafe', 'до 25 Вт', 'до 15 Вт', True),
            ('Батарея', 'до 33 ч видео', 'до 29 ч видео', True),
            ('USB-C', 'USB 3 (10 Гбит/с)', 'USB 3 (10 Гбит/с)', False),
        ],
    },
    'iphone-16-pro': {
        'compareTitle': 'iPhone 15 Pro',
        'specs': [
            ('Процессор', 'Apple A18 Pro'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '6,3" Super Retina XDR OLED, 120 Гц ProMotion'),
            ('Основная камера', '48 МП + 48 МП Ultra Wide + 12 МП (5× зум)'),
            ('Корпус', 'Титан'),
            ('Разъём', 'USB-C (USB 3)'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
        ],
        'compare': [
            ('Процессор', 'A18 Pro', 'A17 Pro', True),
            ('Диагональ экрана', '6,3"', '6,1"', True),
            ('Оптический зум', '5×', '3×', True),
            ('Ultra Wide', '48 МП', '12 МП', True),
            ('Camera Control', 'Да', 'Нет', True),
            ('Apple Intelligence', 'Да', 'Ограниченная', True),
            ('MagSafe', 'до 25 Вт', 'до 15 Вт', True),
        ],
    },
    'iphone-16-plus': {
        'compareTitle': 'iPhone 15 Plus',
        'specs': [
            ('Процессор', 'Apple A18'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '6,7" Super Retina XDR OLED'),
            ('Основная камера', '48 МП + 12 МП Ultra Wide'),
            ('Фронтальная', '12 МП TrueDepth'),
            ('Батарея', 'до 27 часов видео'),
            ('Разъём', 'USB-C'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
            ('MagSafe', 'До 25 Вт'),
        ],
        'compare': [
            ('Процессор', 'A18', 'A16 Bionic', True),
            ('Оперативная память', '8 ГБ', '6 ГБ', True),
            ('Apple Intelligence', 'Да', 'Нет', True),
            ('Camera Control', 'Да', 'Нет', True),
            ('Action Button', 'Да', 'Нет', True),
            ('Макросъёмка', 'Да', 'Нет', True),
            ('MagSafe', 'до 25 Вт', 'до 15 Вт', True),
        ],
    },
    'iphone-16': {
        'compareTitle': 'iPhone 15',
        'specs': [
            ('Процессор', 'Apple A18'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '6,1" Super Retina XDR OLED'),
            ('Основная камера', '48 МП + 12 МП Ultra Wide'),
            ('Фронтальная', '12 МП TrueDepth'),
            ('Батарея', 'до 22 часов видео'),
            ('Разъём', 'USB-C'),
            ('Apple Intelligence', 'Да'),
            ('Camera Control', 'Да'),
            ('Action Button', 'Да'),
        ],
        'compare': [
            ('Процессор', 'A18', 'A16 Bionic', True),
            ('Оперативная память', '8 ГБ', '6 ГБ', True),
            ('Apple Intelligence', 'Да', 'Нет', True),
            ('Camera Control', 'Да', 'Нет', True),
            ('Action Button', 'Да', 'Нет', True),
            ('Макросъёмка', 'Да', 'Нет', True),
            ('MagSafe', 'до 25 Вт', 'до 15 Вт', True),
        ],
    },
    'iphone-16e': {
        'compareTitle': 'iPhone SE (3-е поколение)',
        'specs': [
            ('Процессор', 'Apple A18 (с 4-ядерным GPU)'),
            ('Дисплей', '6,1" Super Retina XDR OLED'),
            ('Основная камера', '48 МП Fusion (2-в-1: основная + 2× кроп)'),
            ('Разъём', 'USB-C'),
            ('Face ID', 'Да'),
            ('Apple Intelligence', 'Да'),
            ('Action Button', 'Да'),
            ('5G', 'Да (Apple C1 модем)'),
            ('Батарея', 'до 26 часов видео'),
        ],
        'compare': [
            ('Процессор', 'A18', 'A15 Bionic', True),
            ('Дисплей', '6,1" OLED', '4,7" LCD', True),
            ('Разъём', 'USB-C', 'Lightning', True),
            ('Разблокировка', 'Face ID', 'Touch ID', True),
            ('Камера', '48 МП Fusion', '12 МП', True),
            ('Apple Intelligence', 'Да', 'Нет', True),
            ('Батарея', 'до 26 ч', 'до 15 ч', True),
            ('5G', 'Да', 'Нет', True),
        ],
    },
    'iphone-15': {
        'compareTitle': 'iPhone 14',
        'specs': [
            ('Процессор', 'Apple A16 Bionic'),
            ('Оперативная память', '6 ГБ'),
            ('Дисплей', '6,1" Super Retina XDR OLED, Dynamic Island'),
            ('Основная камера', '48 МП + 12 МП Ultra Wide'),
            ('Разъём', 'USB-C'),
            ('Dynamic Island', 'Да'),
            ('5G', 'Да'),
        ],
        'compare': [
            ('Процессор', 'A16 Bionic', 'A15 Bionic', True),
            ('Dynamic Island', 'Да', 'Нет (вырез)', True),
            ('Разъём', 'USB-C', 'Lightning', True),
            ('Основная камера', '48 МП', '12 МП', True),
            ('Яркость дисплея', '2000 нит (пик)', '1200 нит', True),
            ('5G', 'Да', 'Да', False),
        ],
    },

    # ═══ iPad ═══════════════════════════════════════════════════════════════
    'ipad-11-2025': {
        'compareTitle': 'iPad 10 (2022)',
        'specs': [
            ('Процессор', 'Apple A16 Bionic'),
            ('Оперативная память', '6 ГБ'),
            ('Дисплей', '11" Liquid Retina, 60 Гц'),
            ('Камера тыловая', '12 МП'),
            ('Камера фронтальная', '12 МП Ultra Wide (ландшафтная)'),
            ('Разъём', 'USB-C'),
            ('Apple Pencil', 'USB-C, 1-го поколения (с адаптером)'),
            ('Батарея', 'до 10 часов Wi-Fi'),
            ('Touch ID', 'В кнопке питания'),
            ('Wi-Fi', 'Wi-Fi 6'),
        ],
        'compare': [
            ('Процессор', 'A16 Bionic', 'A14 Bionic', True),
            ('Оперативная память', '6 ГБ', '4 ГБ', True),
            ('Накопитель старт', '128 ГБ', '64 ГБ', True),
            ('Apple Intelligence', 'Нет', 'Нет', False),
            ('Apple Pencil Pro', 'Нет', 'Нет', False),
            ('Wi-Fi', 'Wi-Fi 6', 'Wi-Fi 6', False),
        ],
    },
    'ipad-air-11-m3': {
        'compareTitle': 'iPad Air 11 M2',
        'specs': [
            ('Процессор', 'Apple M3'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '11" Liquid Retina, True Tone, P3, антибликовое покрытие'),
            ('Камера тыловая', '12 МП широкоугольная'),
            ('Камера фронтальная', '12 МП Ultra Wide (ландшафтная)'),
            ('Разъём', 'USB-C'),
            ('Apple Pencil', 'Pro + USB-C'),
            ('Apple Intelligence', 'Да'),
            ('Wi-Fi', 'Wi-Fi 6E'),
            ('Touch ID', 'В кнопке питания'),
        ],
        'compare': [
            ('Процессор', 'M3', 'M2', True),
            ('Ядер GPU', '9', '9', False),
            ('Hardware Ray Tracing', 'Да', 'Нет', True),
            ('Dynamic Caching', 'Да', 'Нет', True),
            ('AV1 декодирование', 'Да', 'Да', False),
            ('Apple Intelligence', 'Да', 'Да', False),
            ('Дисплей', '11" Liquid Retina', '11" Liquid Retina', False),
        ],
    },
    'ipad-air-11-m4': {
        'compareTitle': 'iPad Air 11 M3',
        'specs': [
            ('Процессор', 'Apple M4'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '11" Liquid Retina'),
            ('Камера фронтальная', 'Center Stage 12 МП'),
            ('Разъём', 'USB-C'),
            ('Apple Pencil', 'Pro + USB-C'),
            ('Apple Intelligence', 'Да'),
            ('Wi-Fi', 'Wi-Fi 6E'),
        ],
        'compare': [
            ('Процессор', 'M4', 'M3', True),
            ('Neural Engine', 'Новый (M4)', 'M3 Neural Engine', True),
            ('Apple Intelligence', 'Да', 'Да', False),
            ('Дисплей', '11" Liquid Retina', '11" Liquid Retina', False),
        ],
    },
    'ipad-pro-11': {
        'compareTitle': 'iPad Pro 11 M4',
        'specs': [
            ('Процессор', 'Apple M5'),
            ('Оперативная память', '8 / 16 ГБ'),
            ('Дисплей', '11" Ultra Retina XDR (Tandem OLED), 120 Гц ProMotion'),
            ('Яркость', '1000 нит (SDR), 1600 нит (HDR)'),
            ('Камера', '12 МП + LiDAR'),
            ('Камера фронтальная', 'Center Stage 12 МП (ландшафтная)'),
            ('Разъём', 'Thunderbolt / USB-C'),
            ('Apple Pencil', 'Pro + USB-C'),
            ('Face ID', 'Да'),
            ('Apple Intelligence', 'Да'),
            ('Wi-Fi', 'Wi-Fi 6E'),
        ],
        'compare': [
            ('Процессор', 'M5', 'M4', True),
            ('Hardware Ray Tracing', 'Да (улучшенный)', 'Да', True),
            ('Дисплей', 'Tandem OLED 120 Гц', 'Tandem OLED 120 Гц', False),
            ('Thunderbolt', 'Да', 'Да', False),
            ('Face ID', 'Да', 'Да', False),
        ],
    },
    'ipad-pro-13': {
        'compareTitle': 'iPad Pro 13 M4',
        'specs': [
            ('Процессор', 'Apple M5'),
            ('Оперативная память', '8 / 16 ГБ'),
            ('Дисплей', '13" Ultra Retina XDR (Tandem OLED), 120 Гц ProMotion'),
            ('Камера', '12 МП + LiDAR'),
            ('Разъём', 'Thunderbolt / USB-C'),
            ('Apple Pencil', 'Pro + USB-C'),
            ('Face ID', 'Да'),
            ('Apple Intelligence', 'Да'),
        ],
        'compare': [
            ('Процессор', 'M5', 'M4', True),
            ('Диагональ', '13"', '13"', False),
            ('Дисплей', 'Tandem OLED', 'Tandem OLED', False),
            ('Thunderbolt', 'Да', 'Да', False),
        ],
    },

    # ═══ MacBook / Mac ═══════════════════════════════════════════════════════
    'mac-mini-m4': {
        'compareTitle': 'Mac mini M2',
        'specs': [
            ('Процессор', 'Apple M4 (10 ядер CPU, 10 ядер GPU)'),
            ('Размер', '12,7 × 12,7 × 5,0 см'),
            ('Порты', '2× Thunderbolt 4, 3× USB-C, HDMI, Gigabit Ethernet, 3,5 мм jack'),
            ('Wi-Fi', 'Wi-Fi 6E'),
            ('Bluetooth', '5.3'),
            ('Apple Intelligence', 'Да'),
            ('Hardware Ray Tracing', 'Да'),
        ],
        'compare': [
            ('Процессор', 'M4', 'M2', True),
            ('Ядер CPU', '10', '8', True),
            ('Размер корпуса', '12,7 × 12,7 см', '19,7 × 19,7 см', True),
            ('Thunderbolt', 'TB4', 'TB4', False),
            ('Hardware Ray Tracing', 'Да', 'Нет', True),
            ('Apple Intelligence', 'Да', 'Нет', True),
        ],
    },
    'macbook-neo': {
        'compareTitle': 'MacBook Air 13" M2',
        'specs': [
            ('Процессор', 'MediaTek Kompanio Ultra'),
            ('Дисплей', '13" Liquid Retina'),
            ('Память', '8 ГБ RAM / 256 ГБ SSD'),
            ('Батарея', 'до 18 часов'),
            ('ОС', 'Neo OS (не macOS)'),
            ('Порты', 'USB-C, USB-A, HDMI'),
            ('Вес', 'около 1,3 кг'),
        ],
        'compare': [
            ('ОС', 'Neo OS', 'macOS', False),
            ('Процессор', 'MediaTek Kompanio Ultra', 'Apple M2', False),
            ('Цена старт', 'от 64 500 ₽', 'от 77 500 ₽', True),
            ('macOS и Mac App Store', 'Нет', 'Да', False),
            ('Apple Intelligence', 'Нет', 'Нет', False),
        ],
    },
    'macbook-air-13-m2': {
        'compareTitle': 'MacBook Air 13" M1',
        'specs': [
            ('Процессор', 'Apple M2 (8 ядер CPU, 8 / 10 ядер GPU)'),
            ('Дисплей', '13,6" Liquid Retina, True Tone, P3'),
            ('Память', 'от 8 ГБ / 256 ГБ SSD'),
            ('Батарея', 'до 18 часов'),
            ('Разъёмы', '2× Thunderbolt / USB-C, MagSafe 3, 3,5 мм jack'),
            ('Камера', '1080p FaceTime HD'),
            ('Вес', '1,24 кг'),
        ],
        'compare': [
            ('Процессор', 'M2', 'M1', True),
            ('MagSafe 3', 'Да', 'Нет', True),
            ('Дисплей', '13,6" (2560×1664)', '13,3" (2560×1600)', True),
            ('Дизайн', 'Плоский', 'Клин', True),
            ('Камера', '1080p', '720p', True),
            ('Динамики', '4 динамика', '2 динамика', True),
        ],
    },
    'macbook-air-13-m5': {
        'compareTitle': 'MacBook Air 13" M4',
        'specs': [
            ('Процессор', 'Apple M5'),
            ('Дисплей', '13,6" Liquid Retina'),
            ('Память', 'от 16 ГБ RAM / 512 ГБ SSD'),
            ('Батарея', 'до 18 часов'),
            ('Apple Intelligence', 'Да'),
            ('MagSafe 3', 'Да'),
            ('Разъёмы', '2× Thunderbolt 4, MagSafe 3'),
        ],
        'compare': [
            ('Процессор', 'M5', 'M4', True),
            ('Neural Engine', 'Новое поколение', 'M4 Neural Engine', True),
            ('Apple Intelligence', 'Да', 'Да', False),
            ('Память минимум', '16 ГБ', '16 ГБ', False),
        ],
    },
    'macbook-air-15-m4': {
        'compareTitle': 'MacBook Air 15" M3',
        'specs': [
            ('Процессор', 'Apple M4'),
            ('Дисплей', '15,3" Liquid Retina'),
            ('Память', 'от 16 ГБ RAM / от 256 ГБ SSD'),
            ('Батарея', 'до 18 часов'),
            ('Apple Intelligence', 'Да'),
            ('Динамики', '6 динамиков с Spatial Audio'),
        ],
        'compare': [
            ('Процессор', 'M4', 'M3', True),
            ('Apple Intelligence', 'Да', 'Да', False),
            ('Память минимум', '16 ГБ', '8 ГБ', True),
        ],
    },
    'macbook-air-15-m5': {
        'compareTitle': 'MacBook Air 15" M4',
        'specs': [
            ('Процессор', 'Apple M5'),
            ('Дисплей', '15,3" Liquid Retina'),
            ('Память', '16 ГБ RAM / 512 ГБ SSD'),
            ('Apple Intelligence', 'Да'),
        ],
        'compare': [
            ('Процессор', 'M5', 'M4', True),
            ('Neural Engine', 'Новое поколение', 'M4 NE', True),
        ],
    },
    'macbook-pro-14-m5': {
        'compareTitle': 'MacBook Pro 14" M4',
        'specs': [
            ('Процессор', 'Apple M5'),
            ('Дисплей', '14,2" Liquid Retina XDR, ProMotion 120 Гц'),
            ('Память', 'от 16 ГБ RAM / от 512 ГБ SSD'),
            ('Разъёмы', '3× Thunderbolt 4, HDMI, SD, MagSafe 3, 3,5 мм'),
            ('Apple Intelligence', 'Да'),
            ('Батарея', 'до 24 часов'),
        ],
        'compare': [
            ('Процессор', 'M5', 'M4', True),
            ('Hardware Ray Tracing', 'Да (улучшенный)', 'Да', True),
            ('Дисплей', '120 Гц ProMotion', '120 Гц ProMotion', False),
            ('Разъёмы', 'Thunderbolt 4', 'Thunderbolt 4', False),
        ],
    },
    'macbook-pro-14-m5-pro': {
        'compareTitle': 'MacBook Pro 14" M4 Pro',
        'specs': [
            ('Процессор', 'Apple M5 Pro'),
            ('Дисплей', '14,2" Liquid Retina XDR, ProMotion 120 Гц'),
            ('Память', '24 ГБ RAM / 1 ТБ SSD'),
            ('Разъёмы', '3× Thunderbolt 5, HDMI, SD, MagSafe 3, 3,5 мм'),
            ('Apple Intelligence', 'Да'),
            ('Батарея', 'до 22 часов'),
        ],
        'compare': [
            ('Процессор', 'M5 Pro', 'M4 Pro', True),
            ('Thunderbolt', 'TB5 (120 Гбит/с)', 'TB5 (120 Гбит/с)', False),
            ('Hardware Ray Tracing', 'Да (улучшенный)', 'Да', True),
            ('Память старт', '24 ГБ', '24 ГБ', False),
        ],
    },

    # ═══ Apple Watch / Galaxy Watch ══════════════════════════════════════════
    'apple-watch-ultra-3': {
        'compareTitle': 'Apple Watch Ultra 2',
        'specs': [
            ('Корпус', 'Титан 49 мм'),
            ('Дисплей', 'Always-On Retina LTPO3, до 3000 нит'),
            ('Чип', 'S10 SiP'),
            ('Батарея', 'до 42 часов (72 часа в режиме энергосбережения)'),
            ('Защита', '100 м водозащиты (ISO 22810), IP6X'),
            ('Подключение', '5G, LTE, Wi-Fi, спутниковая связь'),
            ('Датчики', 'ECG, SpO2, температура, High/Low HR alerts'),
            ('Apple Intelligence', 'Нет (Siri on-device)'),
        ],
        'compare': [
            ('Чип', 'S10 SiP', 'S9 SiP', True),
            ('Батарея', 'до 42 ч', 'до 36 ч', True),
            ('Спутниковая связь', 'Да', 'Нет', True),
            ('5G', 'Да', 'Нет', True),
            ('Яркость', '3000 нит', '3000 нит', False),
            ('Корпус', 'Титан 49 мм', 'Титан 49 мм', False),
            ('Двойное нажатие', 'Да', 'Да', False),
        ],
    },
    'apple-watch-ultra-2': {
        'compareTitle': 'Apple Watch Ultra',
        'specs': [
            ('Корпус', 'Титан 49 мм'),
            ('Дисплей', 'Always-On Retina, до 3000 нит'),
            ('Чип', 'S9 SiP'),
            ('Батарея', 'до 36 часов (72 часа в режиме энергосбережения)'),
            ('Защита', '100 м водозащиты, IP6X'),
            ('Датчики', 'ECG, SpO2, температура'),
            ('Двойное нажатие', 'Да'),
        ],
        'compare': [
            ('Чип', 'S9 SiP', 'S8 SiP', True),
            ('Двойное нажатие', 'Да', 'Нет', True),
            ('Яркость дисплея', '3000 нит', '2000 нит', True),
            ('On-device Siri', 'Да', 'Нет', True),
            ('Батарея', 'до 36 ч', 'до 36 ч', False),
        ],
    },
    'apple-watch-ultra-2025': {
        'compareTitle': 'Apple Watch Ultra 2',
        'specs': [
            ('Корпус', 'Титан 49 мм'),
            ('Дисплей', 'Always-On Retina'),
            ('Батарея', 'до 36 часов'),
            ('Защита', '100 м водозащиты'),
        ],
        'compare': [
            ('Корпус', 'Титан 49 мм', 'Титан 49 мм', False),
            ('Дисплей', 'Always-On Retina', 'Always-On Retina', False),
        ],
    },
    'apple-watch-series-11': {
        'compareTitle': 'Apple Watch Series 10',
        'specs': [
            ('Корпус', 'Алюминий / титан, 42 мм и 46 мм'),
            ('Дисплей', 'Always-On Retina LTPO3'),
            ('Чип', 'S11 SiP'),
            ('Батарея', 'до 24 часов'),
            ('Защита', 'IP6X, 50 м водозащиты'),
            ('Датчики', 'ECG, SpO2, температура, пульс, сон'),
            ('Подключение', '5G опц., Wi-Fi 4'),
        ],
        'compare': [
            ('Чип', 'S11 SiP', 'S10 SiP', True),
            ('Батарея', 'до 24 ч', 'до 18 ч', True),
            ('5G', 'Да (опц.)', 'Нет', True),
            ('Измерение давления', 'Да (Hypertension alerts)', 'Нет', True),
            ('Apple Intelligence', 'Частично', 'Нет', True),
            ('Размеры', '42 и 46 мм', '42 и 46 мм', False),
        ],
    },
    'apple-watch-series-10': {
        'compareTitle': 'Apple Watch Series 9',
        'specs': [
            ('Корпус', 'Алюминий, 42 мм и 46 мм'),
            ('Дисплей', 'Always-On Retina, шире чем у Series 9'),
            ('Чип', 'S10 SiP'),
            ('Батарея', 'до 18 часов'),
            ('Толщина', '9,7 мм (тоньше Series 9)'),
        ],
        'compare': [
            ('Чип', 'S10 SiP', 'S9 SiP', True),
            ('Размер экрана', '42/46 мм', '41/45 мм', True),
            ('Толщина', '9,7 мм', '10,7 мм', True),
            ('Угол обзора', 'Шире (новый OLED)', 'Стандарт', True),
            ('Спящие измерения', 'Да', 'Нет', True),
        ],
    },
    'apple-watch-se-3': {
        'compareTitle': 'Apple Watch SE 2',
        'specs': [
            ('Корпус', 'Алюминий, 40 мм и 44 мм'),
            ('Дисплей', 'Retina (без Always-On)'),
            ('Чип', 'S10 SiP'),
            ('Apple Intelligence', 'Частично'),
            ('Датчики', 'Пульс, акселерометр, гироскоп, SOS'),
        ],
        'compare': [
            ('Чип', 'S10 SiP', 'S8 SiP', True),
            ('Apple Intelligence', 'Да (частично)', 'Нет', True),
            ('Siri on-device', 'Да', 'Нет', True),
            ('Always-On дисплей', 'Нет', 'Нет', False),
            ('ECG', 'Нет', 'Нет', False),
        ],
    },
    'apple-watch-se-2': {
        'compareTitle': 'Apple Watch SE (1-е поколение)',
        'specs': [
            ('Корпус', 'Алюминий, 40 мм и 44 мм'),
            ('Чип', 'S8 SiP'),
            ('Батарея', 'до 18 часов'),
            ('Датчики', 'Пульс, акселерометр, SOS'),
        ],
        'compare': [
            ('Чип', 'S8 SiP', 'S5 SiP', True),
            ('Crash Detection', 'Да', 'Нет', True),
            ('Always-On', 'Нет', 'Нет', False),
        ],
    },
    'samsung-galaxy-watch-8': {
        'compareTitle': 'Galaxy Watch 7',
        'specs': [
            ('Корпус', '40 мм / 44 мм'),
            ('ОС', 'Wear OS 6 / One UI Watch 8'),
            ('Чип', 'Exynos W1000 (3 нм)'),
            ('Датчики', 'BioActive (пульс, ECG, SpO2, давление, температура)'),
            ('Защита', '5 ATM, IP68'),
            ('GPS', 'Двухчастотный L1+L5'),
        ],
        'compare': [
            ('Чип', 'Exynos W1000', 'Exynos W1000', False),
            ('ОС', 'Wear OS 6', 'Wear OS 5', True),
            ('Galaxy AI', 'Да', 'Ограниченно', True),
            ('Дизайн', 'Обновлённый тонкий', 'Классический', True),
        ],
    },
    'samsung-galaxy-watch-classic-8': {
        'compareTitle': 'Galaxy Watch 7 Classic',
        'specs': [
            ('Корпус', 'Нержавеющая сталь 46 мм'),
            ('Безель', 'Физический поворотный'),
            ('ОС', 'Wear OS 6 / One UI Watch 8'),
            ('Чип', 'Exynos W1000'),
            ('Датчики', 'BioActive полный набор'),
        ],
        'compare': [
            ('Корпус', 'Сталь 46 мм', 'Сталь 47 мм', False),
            ('Поворотный безель', 'Да', 'Да', False),
            ('Galaxy AI', 'Да', 'Нет', True),
        ],
    },

    # ═══ AirPods / наушники ══════════════════════════════════════════════════
    'airpods-max': {
        'compareTitle': 'AirPods Max (2020)',
        'specs': [
            ('Тип', 'Полноразмерные накладные (over-ear)'),
            ('Чип', 'H1 (в каждой чашке)'),
            ('Активное шумоподавление', 'Да'),
            ('Adaptive EQ', 'Да'),
            ('Transparency Mode', 'Да'),
            ('Personalized Spatial Audio', 'Да'),
            ('Разъём кейса', 'USB-C'),
            ('Батарея', 'до 20 часов с ANC'),
            ('Вес', '384 г'),
        ],
        'compare': [
            ('Разъём', 'USB-C', 'Lightning', True),
            ('Новые цвета', 'Purple, Orange, Starlight, Midnight, Blue', '5 цветов (старые)', False),
            ('Чип', 'H1', 'H1', False),
            ('Батарея', 'до 20 ч', 'до 20 ч', False),
            ('Активное шумоподавление', 'Да', 'Да', False),
        ],
    },
    'airpods-pro-3': {
        'compareTitle': 'AirPods Pro 2',
        'specs': [
            ('Чип', 'H3'),
            ('Активное шумоподавление', 'Да (2-кратное улучшение vs Pro 2)'),
            ('Adaptive Audio', 'Да'),
            ('Conversation Awareness', 'Да'),
            ('Heart Rate Sensor', 'Да (встроенный пульсометр)'),
            ('Live Translation', 'Да'),
            ('Защита', 'IP57 (наушники и кейс)'),
            ('Кейс', 'USB-C + MagSafe'),
            ('Батарея', 'до 8 ч воспроизведения, до 10 ч без ANC'),
        ],
        'compare': [
            ('Чип', 'H3', 'H2', True),
            ('ANC', '2× лучше', 'Стандарт', True),
            ('Пульсометр', 'Да', 'Нет', True),
            ('Live Translation', 'Да', 'Нет', True),
            ('Защита', 'IP57', 'IP54', True),
            ('Батарея', 'до 8 ч', 'до 6 ч', True),
        ],
    },
    'airpods-pro-2': {
        'compareTitle': 'AirPods Pro (1-е поколение)',
        'specs': [
            ('Чип', 'H2'),
            ('Активное шумоподавление', 'Да (2× vs 1-го поколения)'),
            ('Adaptive Audio', 'Да'),
            ('Personalized Spatial Audio', 'Да'),
            ('Conversation Awareness', 'Да'),
            ('Hearing Aid', 'Да'),
            ('Защита', 'IP54'),
            ('Кейс', 'USB-C + MagSafe'),
        ],
        'compare': [
            ('Чип', 'H2', 'H1', True),
            ('ANC', '2× лучше', 'Стандарт', True),
            ('Adaptive Audio', 'Да', 'Нет', True),
            ('Разъём кейса', 'USB-C', 'Lightning', True),
            ('Функция слухового аппарата', 'Да', 'Нет', True),
        ],
    },
    'airpods-4': {
        'compareTitle': 'AirPods 3',
        'specs': [
            ('Тип', 'Открытые (без амбушюр)'),
            ('Чип', 'H2'),
            ('Активное шумоподавление', 'Только в версии с ANC'),
            ('Personalized Spatial Audio', 'Да'),
            ('Кейс', 'USB-C'),
            ('Батарея', 'до 5 ч воспроизведения (до 30 ч с кейсом)'),
        ],
        'compare': [
            ('Чип', 'H2', 'H1', True),
            ('ANC версия', 'Да (опц.)', 'Нет', True),
            ('Кейс с колонкой', 'Да (версия с ANC)', 'Нет', True),
            ('Разъём', 'USB-C', 'Lightning', True),
        ],
    },
    'galaxy-buds': {
        'compareTitle': 'Galaxy Buds 2 Pro',
        'specs': [
            ('Модели', 'Buds 3 / Buds 3 Pro / Buds 3 FE'),
            ('Активное шумоподавление', 'Да (Pro — улучшенное)'),
            ('Galaxy AI', 'Live Translate, Interpreter'),
            ('Водозащита', 'IP57'),
            ('Батарея', 'до 6 ч с ANC'),
        ],
        'compare': [
            ('Galaxy AI', 'Да (Live Translate)', 'Нет', True),
            ('Дизайн', 'Stem-форма', 'Bean-форма', False),
            ('ANC', 'Улучшенный', 'Стандарт', True),
        ],
    },
    'marshall-headphones': {
        'compareTitle': 'Marshall Monitor II',
        'specs': [
            ('Модели', 'Major IV, Major V'),
            ('Тип', 'Накладные, проводные/беспроводные'),
            ('Подключение', 'Bluetooth 5.3'),
            ('Батарея', 'до 80 часов (Major V)'),
            ('Управление', 'Multi-directional control knob'),
            ('Дизайн', 'Легендарный Marshall стиль'),
        ],
        'compare': [
            ('Время работы', 'до 80 ч (Major V)', 'до 30 ч', True),
            ('Bluetooth', '5.3', '5.2', True),
            ('Быстрая зарядка', '15 мин = 15 ч', 'Нет', True),
        ],
    },

    # ═══ Android ═══════════════════════════════════════════════════════════════
    'samsung-galaxy-s26-ultra': {
        'compareTitle': 'Samsung Galaxy S25 Ultra',
        'specs': [
            ('Процессор', 'Snapdragon 8 Elite Gen 5 for Galaxy'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,9" QHD+ Dynamic AMOLED 2X, 120 Гц'),
            ('Основная камера', '200 МП + 50 МП перископ + 50 МП телефото + 50 МП Ultra Wide'),
            ('S Pen', 'Да (с Bluetooth)'),
            ('Батарея', '5000 мАч'),
            ('Зарядка', '45 Вт проводная, 15 Вт беспроводная'),
            ('Корпус', 'Титан (Grade 5)'),
            ('Galaxy AI', 'Да (расширенный)'),
            ('ОС', 'Android 16 / One UI 8'),
        ],
        'compare': [
            ('Процессор', 'SD 8 Elite Gen 5', 'SD 8 Elite', True),
            ('Оперативная память', '12 ГБ', '12 ГБ', False),
            ('Ultra Wide', '50 МП', '50 МП', False),
            ('Основная камера', '200 МП', '200 МП', False),
            ('Galaxy AI', 'Расширенный', 'Базовый', True),
            ('Зарядка', '45 Вт', '45 Вт', False),
            ('S Pen', 'Да', 'Да', False),
        ],
    },
    'samsung-galaxy-s26-plus': {
        'compareTitle': 'Samsung Galaxy S25+',
        'specs': [
            ('Процессор', 'Snapdragon 8 Elite Gen 5'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,7" QHD+ Dynamic AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 50 МП Ultra Wide + 12 МП'),
            ('Батарея', '4900 мАч'),
            ('Galaxy AI', 'Да'),
        ],
        'compare': [
            ('Процессор', 'SD 8 Elite Gen 5', 'SD 8 Elite', True),
            ('Оперативная память', '12 ГБ', '12 ГБ', False),
            ('Дисплей', '6,7" QHD+', '6,7" QHD+', False),
        ],
    },
    'samsung-galaxy-s26': {
        'compareTitle': 'Samsung Galaxy S25',
        'specs': [
            ('Процессор', 'Snapdragon 8 Elite Gen 5'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,2" FHD+ Dynamic AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 50 МП Ultra Wide + 12 МП'),
            ('Galaxy AI', 'Да'),
        ],
        'compare': [
            ('Процессор', 'SD 8 Elite Gen 5', 'SD 8 Elite', True),
            ('Galaxy AI', 'Расширенный', 'Базовый', True),
        ],
    },
    'samsung-galaxy-s25-ultra': {
        'compareTitle': 'Samsung Galaxy S24 Ultra',
        'specs': [
            ('Процессор', 'Snapdragon 8 Elite for Galaxy'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,9" QHD+ Dynamic AMOLED, 120 Гц'),
            ('Камеры', '200 МП + 50 МП + 50 МП + 10 МП'),
            ('S Pen', 'Да (без Bluetooth)'),
            ('Батарея', '5000 мАч'),
            ('Корпус', 'Титан'),
            ('Galaxy AI', 'Да'),
        ],
        'compare': [
            ('Процессор', 'SD 8 Elite', 'SD 8 Gen 3', True),
            ('S Pen Bluetooth', 'Нет', 'Да', False),
            ('Ultra Wide', '50 МП', '12 МП', True),
            ('Корпус', 'Титан', 'Титан', False),
            ('Galaxy AI', 'Расширенный', 'Базовый', True),
        ],
    },
    'samsung-galaxy-s25-fe': {
        'compareTitle': 'Samsung Galaxy S24 FE',
        'specs': [
            ('Процессор', 'Exynos 2400e'),
            ('Оперативная память', '8 ГБ'),
            ('Дисплей', '6,7" FHD+ Dynamic AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 12 МП Ultra Wide + 8 МП'),
            ('Батарея', '4900 мАч'),
            ('Galaxy AI', 'Да'),
        ],
        'compare': [
            ('Процессор', 'Exynos 2400e', 'Exynos 2400e', False),
            ('Galaxy AI', 'Да', 'Да', False),
            ('Память старт', '8 ГБ', '8 ГБ', False),
        ],
    },
    'samsung-galaxy-s25': {
        'compareTitle': 'Samsung Galaxy S24',
        'specs': [
            ('Процессор', 'Snapdragon 8 Elite for Galaxy'),
            ('Оперативная память', '12 ГБ'),
            ('Дисплей', '6,2" FHD+ Dynamic AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 12 МП Ultra Wide + 10 МП'),
            ('Galaxy AI', 'Да'),
        ],
        'compare': [
            ('Процессор', 'SD 8 Elite', 'SD 8 Gen 3', True),
            ('Оперативная память', '12 ГБ', '8 ГБ', True),
            ('Galaxy AI', 'Да', 'Да', False),
        ],
    },
    'samsung-galaxy-a56': {
        'compareTitle': 'Samsung Galaxy A55',
        'specs': [
            ('Процессор', 'Exynos 1580'),
            ('Оперативная память', '8 / 12 ГБ'),
            ('Дисплей', '6,7" FHD+ Super AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 12 МП Ultra Wide + 5 МП'),
            ('Батарея', '5000 мАч'),
            ('Зарядка', '45 Вт'),
            ('Защита', 'IP67'),
        ],
        'compare': [
            ('Процессор', 'Exynos 1580', 'Exynos 1480', True),
            ('Зарядка', '45 Вт', '25 Вт', True),
            ('Обновления ОС', '6 лет', '5 лет', True),
            ('Защита', 'IP67', 'IP67', False),
        ],
    },
    'samsung-galaxy-a36': {
        'compareTitle': 'Samsung Galaxy A35',
        'specs': [
            ('Процессор', 'Snapdragon 6 Gen 3'),
            ('Дисплей', '6,7" FHD+ Super AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 8 МП + 5 МП'),
            ('Батарея', '5000 мАч'),
            ('Зарядка', '45 Вт'),
            ('Защита', 'IP67'),
        ],
        'compare': [
            ('Процессор', 'SD 6 Gen 3', 'Exynos 1380', True),
            ('Зарядка', '45 Вт', '25 Вт', True),
            ('Обновления', '6 лет', '5 лет', True),
        ],
    },
    'samsung-galaxy-a26': {
        'compareTitle': 'Samsung Galaxy A25',
        'specs': [
            ('Процессор', 'Exynos 1380'),
            ('Дисплей', '6,5" FHD+ Super AMOLED, 120 Гц'),
            ('Камеры', '50 МП + 8 МП + 2 МП'),
            ('Защита', 'IP67'),
        ],
        'compare': [
            ('Защита', 'IP67', 'Нет', True),
            ('Дисплей', '6,5" AMOLED 120 Гц', '6,5" AMOLED 120 Гц', False),
        ],
    },
    'samsung-galaxy-a17': {
        'compareTitle': 'Samsung Galaxy A16',
        'specs': [
            ('Процессор', 'Exynos 1330'),
            ('Дисплей', '6,7" FHD+ Super AMOLED'),
            ('Камеры', '50 МП + 5 МП + 2 МП'),
            ('Батарея', '5000 мАч'),
        ],
        'compare': [
            ('Обновления ОС', '6 лет', '6 лет', False),
            ('Дисплей', '6,7" AMOLED', '6,7" AMOLED', False),
        ],
    },
    'samsung-galaxy-a07': {
        'compareTitle': 'Samsung Galaxy A06',
        'specs': [
            ('Дисплей', '6,7" PLS LCD'),
            ('Оперативная память', '4 / 6 ГБ'),
            ('Батарея', '5000 мАч'),
            ('Камера', '50 МП'),
        ],
        'compare': [],
    },
    'xiaomi-mi-15t': {
        'compareTitle': 'Xiaomi 14T',
        'specs': [
            ('Процессор', 'MediaTek Dimensity 8400-Ultra'),
            ('Дисплей', '6,83" CrystalRes AMOLED, 144 Гц'),
            ('Камеры', 'Leica: 50 МП + 50 МП телефото + 12 МП Ultra Wide'),
            ('Батарея', '5500 мАч'),
            ('Зарядка', '67 Вт HyperCharge'),
            ('Защита', 'IP68'),
        ],
        'compare': [
            ('Процессор', 'Dimensity 8400-Ultra', 'Dimensity 8300-Ultra', True),
            ('Батарея', '5500 мАч', '5000 мАч', True),
            ('Leica оптика', 'Да', 'Да', False),
        ],
    },
    'xiaomi-redmi-note-15-pro': {
        'compareTitle': 'Redmi Note 14 Pro',
        'specs': [
            ('Процессор', 'MediaTek Dimensity 7400-Ultra'),
            ('Дисплей', '6,83" 1.5K AMOLED, 120 Гц'),
            ('Камеры', '200 МП + 8 МП + 2 МП'),
            ('Батарея', '7000 мАч'),
            ('Зарядка', '45 Вт'),
            ('Защита', 'IP68, IP69'),
        ],
        'compare': [
            ('Батарея', '7000 мАч', '5110 мАч', True),
            ('Защита', 'IP68+IP69', 'IP68', True),
            ('Дисплей', '1.5K 120 Гц', '1.5K 120 Гц', False),
        ],
    },
    'xiaomi-redmi-note-15': {
        'compareTitle': 'Redmi Note 14',
        'specs': [
            ('Процессор', 'Snapdragon 7s Gen 4'),
            ('Дисплей', '6,83" 1.5K AMOLED, 120 Гц'),
            ('Камеры', '108 МП + 2 МП'),
            ('Батарея', '7000 мАч'),
            ('Защита', 'IP68, IP69'),
        ],
        'compare': [
            ('Батарея', '7000 мАч', '5500 мАч', True),
            ('Защита', 'IP68+IP69', 'IP64', True),
        ],
    },
    'xiaomi-redmi-note-14s': {
        'compareTitle': 'Redmi Note 13',
        'specs': [
            ('Дисплей', '6,67" AMOLED, 120 Гц'),
            ('Камера', '108 МП'),
            ('Батарея', '5500 мАч'),
        ],
        'compare': [],
    },
    'xiaomi-redmi-note-14': {
        'compareTitle': 'Redmi Note 13',
        'specs': [
            ('Дисплей', '6,67" AMOLED, 120 Гц'),
            ('Камера', '108 МП'),
            ('Батарея', '5500 мАч'),
        ],
        'compare': [
            ('Защита', 'IP64', 'Нет', True),
        ],
    },
    'meizu-note-21': {
        'compareTitle': 'Meizu Note 20',
        'specs': [
            ('Дисплей', '6,75" AMOLED, 120 Гц'),
            ('Процессор', 'Snapdragon 6 Gen 1'),
            ('Батарея', '5000 мАч'),
        ],
        'compare': [],
    },
}


MODEL_NAMES = {
    # iPhone
    'iphone-17-pro-max': 'iPhone 17 Pro Max', 'iphone-17-pro': 'iPhone 17 Pro', 'iphone-17': 'iPhone 17',
    'iphone-16-pro-max': 'iPhone 16 Pro Max', 'iphone-16-pro': 'iPhone 16 Pro',
    'iphone-16-plus': 'iPhone 16 Plus', 'iphone-16': 'iPhone 16', 'iphone-16e': 'iPhone 16e',
    'iphone-15-pro-max': 'iPhone 15 Pro Max', 'iphone-15-pro': 'iPhone 15 Pro',
    'iphone-15-plus': 'iPhone 15 Plus', 'iphone-15': 'iPhone 15',
    # iPad
    'ipad-11-2025': 'iPad 11 (2025)', 'ipad-air-11-m3': 'iPad Air 11" M3',
    'ipad-air-11-m4': 'iPad Air 11" M4', 'ipad-pro-11': 'iPad Pro 11" M5', 'ipad-pro-13': 'iPad Pro 13" M5',
    # MacBook
    'mac-mini-m4': 'Mac mini M4', 'macbook-neo': 'MacBook Neo',
    'macbook-air-13-m2': 'MacBook Air 13" M2', 'macbook-air-13-m5': 'MacBook Air 13" M5',
    'macbook-air-15-m4': 'MacBook Air 15" M4', 'macbook-air-15-m5': 'MacBook Air 15" M5',
    'macbook-pro-14-m5': 'MacBook Pro 14" M5', 'macbook-pro-14-m5-pro': 'MacBook Pro 14" M5 Pro',
    # Watch
    'apple-watch-ultra-3': 'Apple Watch Ultra 3',
    'apple-watch-ultra-2': 'Apple Watch Ultra 2',
    'apple-watch-ultra-2025': 'Apple Watch Ultra (2025)',
    'apple-watch-series-11': 'Apple Watch Series 11',
    'apple-watch-series-10': 'Apple Watch Series 10',
    'apple-watch-se-3': 'Apple Watch SE 3',
    'apple-watch-se-2': 'Apple Watch SE 2',
    'samsung-galaxy-watch-8': 'Samsung Galaxy Watch 8',
    'samsung-galaxy-watch-classic-8': 'Galaxy Watch Classic 8',
    # AirPods
    'airpods-max':   'AirPods Max',
    'airpods-pro-3': 'AirPods Pro 3',
    'airpods-pro-2': 'AirPods Pro 2 (USB-C)',
    'airpods-4':     'AirPods 4',
    'galaxy-buds':   'Samsung Galaxy Buds',
    'marshall-headphones': 'Наушники Marshall',
    # Android
    'samsung-galaxy-s26-ultra': 'Samsung Galaxy S26 Ultra',
    'samsung-galaxy-s26-plus':  'Samsung Galaxy S26+',
    'samsung-galaxy-s26':       'Samsung Galaxy S26',
    'samsung-galaxy-s25-ultra': 'Samsung Galaxy S25 Ultra',
    'samsung-galaxy-s25-fe':    'Samsung Galaxy S25 FE',
    'samsung-galaxy-s25':       'Samsung Galaxy S25',
    'samsung-galaxy-a56':       'Samsung Galaxy A56',
    'samsung-galaxy-a36':       'Samsung Galaxy A36',
    'samsung-galaxy-a26':       'Samsung Galaxy A26',
    'samsung-galaxy-a17':       'Samsung Galaxy A17',
    'samsung-galaxy-a07':       'Samsung Galaxy A07',
    'xiaomi-mi-15t':            'Xiaomi Mi 15T',
    'xiaomi-redmi-note-15-pro': 'Xiaomi Redmi Note 15 Pro',
    'xiaomi-redmi-note-15':     'Xiaomi Redmi Note 15',
    'xiaomi-redmi-note-14s':    'Xiaomi Redmi Note 14S',
    'xiaomi-redmi-note-14':     'Xiaomi Redmi Note 14',
    'meizu-note-21':            'Meizu Note 21',
    # Dyson
    'dyson-hd18-pro':         'Фен Dyson HD18 Pro',
    'dyson-hd17':             'Фен Dyson HD17',
    'dyson-hd16':             'Фен Dyson HD16',
    'dyson-airwrap-hs09':     'Стайлер Dyson Airwrap HS09 Coanda 2x',
    'dyson-airwrap-hs08':     'Стайлер Dyson Airwrap HS08',
    'dyson-airwrap-hs05':     'Стайлер Dyson Airwrap HS05',
    'dyson-airstrait-ht01':   'Выпрямитель Dyson Airstrait HT01',
    'dyson-wash-g1':          'Пылесос Dyson WashG1',
    'dyson-gen5-detect':      'Пылесос Dyson Gen5 Detect',
    'dyson-v16s':             'Пылесос Dyson V16s Submarine',
    'dyson-v15s':             'Пылесос Dyson V15s Submarine',
    'dyson-v15':              'Пылесос Dyson V15 Detect',
    'dyson-v12s':             'Пылесос Dyson V12s Submarine',
    'dyson-v12':              'Пылесос Dyson V12 Detect Slim',
    'dyson-v8':               'Пылесос Dyson V8',
    # Audio
    'jbl-boombox-4':          'JBL Boombox 4',
    'yandex-station-light-2': 'Яндекс Станция Лайт 2',
    'dji-mic-3':              'DJI Mic 3 (2TX + 1RX)',
    'dji-mic-2':              'DJI Mic 2 (2TX + 1RX)',
    'dji-mic-mini':           'DJI Mic Mini (2TX + 1RX)',
}

CAT_TITLE = {
    'iphone': 'смартфон', 'ipad': 'планшет', 'macbook': 'ноутбук',
    'watch': 'умные часы', 'airpods': 'наушники', 'android': 'смартфон',
    'dyson': 'устройство Dyson', 'audio': 'аудио-устройство',
}

UPSELL_BY_CAT = {
    'iphone': [
        ('charger-20w',   'Блок питания Apple 20W USB-C',  'В коробке iPhone только кабель — рекомендуем оригинальный адаптер.', 2490, '🔌'),
        ('case-silicone', 'Чехол Silicone Case с MagSafe', 'Защита корпуса с первого дня.', 3490, '📱'),
        ('glass',         'Защитное стекло + установка',    'Наклеим за 5 минут в магазине.', 990, '🛡️'),
        ('cardholder',    'Картхолдер MagSafe',             'Держатель для карт на MagSafe.', 1290, '💳'),
    ],
    'ipad': [
        ('apple-pencil',  'Apple Pencil Pro',               'Для рисования и заметок.', 14900, '✏️'),
        ('magic-keyboard','Magic Keyboard для iPad',         'Тачпад, подсветка клавиш, USB-C.', 19900, '⌨️'),
        ('ipad-case',     'Smart Folio чехол',              'Защита экрана + подставка.', 3490, '📱'),
        ('charger-30w',   'Блок питания Apple 30W USB-C',   'Быстрая зарядка для iPad.', 3490, '🔌'),
    ],
    'macbook': [
        ('magic-mouse',   'Magic Mouse',                     'Беспроводная мышь Apple.', 8990, '🖱️'),
        ('magic-keyboard','Magic Keyboard с Touch ID',       'Беспроводная клавиатура с Touch ID.', 12990, '⌨️'),
        ('usb-c-hub',     'USB-C Hub (HDMI + USB-A + SD)',   'Расширение разъёмов для MacBook.', 3490, '🔌'),
        ('macbook-sleeve','Чехол-конверт для MacBook',       'Защита от царапин.', 2490, '💼'),
    ],
    'watch': [
        ('strap-sport',   'Спортивный ремешок',             'Силиконовый, для тренировок.', 1990, '⌚'),
        ('strap-leather', 'Кожаный ремешок',                'Для делового стиля.', 3490, '👔'),
        ('charger',       'Зарядное устройство для Watch',  'Магнитная зарядка.', 2490, '🔌'),
        ('case-watch',    'Защитный кейс с пленкой',        'Защита экрана и корпуса.', 990, '🛡️'),
    ],
    'airpods': [
        ('charger-20w',   'Блок питания Apple 20W USB-C',  'Для быстрой зарядки кейса.', 2490, '🔌'),
        ('case-airpods',  'Чехол для кейса AirPods',       'Силиконовый защитный чехол.', 990, '📦'),
        ('eartips',       'Сменные амбушюры',              'Для AirPods Pro: 4 размера.', 1490, '👂'),
        ('cable-usb-c',   'Кабель USB-C / Lightning',       'Оригинальный 1 м.', 1290, '🔗'),
    ],
    'android': [
        ('charger-25w',   'Блок питания 25W USB-C',         'Быстрая зарядка для Galaxy.', 2490, '🔌'),
        ('case-android',  'Чехол для смартфона',            'Защита корпуса.', 1990, '📱'),
        ('glass-android', 'Защитное стекло + установка',    'Полное закрытие экрана.', 990, '🛡️'),
        ('headphones',    'Беспроводные наушники',          'Galaxy Buds 3 / Marshall.', 9990, '🎧'),
    ],
    'dyson': [
        ('case-dyson',    'Кейс для хранения',              'Защита от пыли и царапин.', 4990, '💼'),
        ('filter-dyson',  'Сменный фильтр',                  'Оригинальный фильтр HEPA.', 2990, '🌀'),
        ('attachment',    'Дополнительная насадка',         'Расширьте функционал.', 3490, '🔧'),
    ],
    'audio': [
        ('case-speaker',  'Чехол для колонки',              'Защита от ударов и пыли.', 1990, '💼'),
        ('cable-audio',   'Аудио-кабель (Aux/USB-C)',       'Качественный кабель 1.5 м.', 990, '🔗'),
        ('charger-fast',  'Зарядное устройство быстрое',    'Для колонок и микрофонов.', 1990, '🔌'),
    ],
}


def generate_category_ts(category, configs):
    lines = [
        f'/** {category}-configs.ts — АВТОГЕНЕРАЦИЯ */',
        'import type { ProductConfig, UpsellItem } from "../product-configs";',
        '',
        'const UPSELL: UpsellItem[] = [',
    ]
    for uid, name, desc, price, emoji in UPSELL_BY_CAT[category]:
        lines.append(f'  {{ id: {ts_str(uid)}, name: {ts_str(name)}, description: {ts_str(desc)}, price: {price}, emoji: {ts_str(emoji)} }},')
    lines.append('];')
    lines.append('')

    for slug, cfg in configs.items():
        meta = MODEL_META.get(slug, {})
        const_name = re.sub(r'[^A-Z0-9]', '_', slug.upper()) + '_CONFIG'
        lines.append(f'const {const_name}: ProductConfig = {{')
        lines.append(f'  slug: {ts_str(slug)},')
        lines.append(f'  category: {ts_str(category)},')
        lines.append('  colors: [')
        for c in cfg['colors']:
            lines.append(f'    {{ id: {ts_str(c["id"])}, name: {ts_str(c["name"])}, hex: {ts_str(c["hex"])}, image: {ts_str(c["image"])} }},')
        lines.append('  ],')
        lines.append('  storage: [')
        for s in cfg['storage']:
            lines.append(f'    {{ id: {ts_str(s["id"])}, label: {ts_str(s["label"])}, available: true }},')
        lines.append('  ],')
        lines.append('  sim: [')
        for s in cfg['sim']:
            lines.append(f'    {{ id: {ts_str(s["id"])}, label: {ts_str(s["label"])}, description: {ts_str(s["description"])} }},')
        lines.append('  ],')
        lines.append('  prices: [')
        for p in cfg['prices']:
            lines.append(f'    {{ storageId: {ts_str(p["storageId"])}, colorId: {ts_str(p["colorId"])}, simId: {ts_str(p["simId"])}, price: {p["price"]} }},')
        lines.append('  ],')
        lines.append(f'  defaultStorage: {ts_str(cfg["defaultStorage"])},')
        lines.append(f'  defaultColor: {ts_str(cfg["defaultColor"])},')
        lines.append(f'  defaultSim: {ts_str(cfg["defaultSim"])},')
        lines.append(f'  priceFrom: {cfg["priceFrom"]},')

        # storageLabel и showSim — по категории
        if category == 'iphone':
            lines.append('  storageLabel: "Объём памяти",')
            lines.append('  showSim: true,')
        elif category == 'ipad':
            lines.append('  storageLabel: "Объём памяти",')
            lines.append(f'  showSim: {"true" if len(cfg["sim"]) > 1 else "false"},')
        elif category == 'macbook':
            lines.append('  storageLabel: "Конфигурация RAM/SSD",')
            lines.append('  showSim: false,')
        elif category == 'watch':
            lines.append('  storageLabel: "Размер корпуса",')
            lines.append(f'  showSim: {"true" if len(cfg["sim"]) > 1 else "false"},')
        elif category == 'android':
            lines.append('  storageLabel: "Память (RAM/Накопитель)",')
            lines.append('  showSim: false,')
        else:  # airpods, dyson, audio
            lines.append('  storageLabel: "Вариант",')
            lines.append('  showSim: false,')

        lines.append('  specs: [')
        for lbl, val in meta.get('specs', []):
            lines.append(f'    {{ label: {ts_str(lbl)}, value: {ts_str(val)} }},')
        lines.append('  ],')
        lines.append(f'  compareTitle: {ts_str(meta.get("compareTitle", ""))},')
        lines.append('  compare: [')
        for lbl, cur, prev, better in meta.get('compare', []):
            lines.append(f'    {{ label: {ts_str(lbl)}, current: {ts_str(cur)}, previous: {ts_str(prev)}, better: {"true" if better else "false"} }},')
        lines.append('  ],')
        lines.append('  upsell: UPSELL,')

        name = MODEL_NAMES.get(slug, slug)
        pf = f'{cfg["priceFrom"]:,}'.replace(',', ' ')
        ct = CAT_TITLE.get(category, 'товар')

        seo_text = (f"{name} — популярный {ct} в магазине ЭПЛ-КОЛЛЕКЦИЯ в Казани. "
                    f"Цена от {pf} ₽. Оригинал с гарантией 1 год, рассрочка 0% на 10 месяцев, "
                    f"бесплатная доставка в день заказа.")
        why_text = (f"В ЭПЛ-КОЛЛЕКЦИЯ каждый {name} проходит проверку перед продажей: "
                    f"тестируем все функции, проверяем серийный номер, "
                    f"активируем и настраиваем устройство прямо в магазине.")

        lines.append(f'  seoH2: {ts_str(f"Купить {name} в Казани")},')
        lines.append(f'  seoText: {ts_str(seo_text)},')
        lines.append(f'  seoH2Why: {ts_str(f"Почему {name} стоит купить у нас?")},')
        lines.append(f'  seoTextWhy: {ts_str(why_text)},')
        if category == 'iphone':
            lines.append(f'  seoH2Sim: {ts_str(f"Какую версию {name} выбрать?")},')
            lines.append(f'  seoTextSim: {ts_str("Nano-SIM + eSIM (Европа/РФ) подходит для большинства операторов. Версия только eSIM — для тех, чей оператор поддерживает eSIM. Уточняйте у менеджера.")},')
        elif category == 'ipad' and any(s['id'] == 'lte' for s in cfg['sim']):
            lines.append(f'  seoH2Sim: {ts_str("Wi-Fi или Wi-Fi + Cellular?")},')
            lines.append(f'  seoTextSim: {ts_str("Wi-Fi — для дома/офиса. Cellular (LTE) — мобильный интернет в любом месте через eSIM.")},')

        lines.append('};')
        lines.append('')

    lines.append('const configs: Record<string, ProductConfig> = {')
    for slug in configs:
        const_name = re.sub(r'[^A-Z0-9]', '_', slug.upper()) + '_CONFIG'
        lines.append(f'  {ts_str(slug)}: {const_name},')
    lines.append('};')
    lines.append('')
    cat_cap = ''.join(w.capitalize() for w in category.split('-'))
    lines.append(f'export function get{cat_cap}Config(slug: string): ProductConfig | undefined {{')
    lines.append('  return configs[slug];')
    lines.append('}')
    lines.append('')
    lines.append(f'export const {category.upper()}_CONFIG_SLUGS = Object.keys(configs);')
    lines.append('')
    return '\n'.join(lines)


def main():
    Path(OUT_DIR).mkdir(parents=True, exist_ok=True)

    # iPhone
    iphone_raw = parse_iphone()
    iphone_order = ['iphone-17-pro-max', 'iphone-17-pro', 'iphone-17',
                    'iphone-16-pro-max', 'iphone-16-pro', 'iphone-16-plus', 'iphone-16', 'iphone-16e',
                    'iphone-15-pro-max', 'iphone-15-pro', 'iphone-15-plus', 'iphone-15']
    iphone_cfgs = OrderedDict()
    for slug in iphone_order:
        if slug in iphone_raw and iphone_raw[slug]:
            iphone_cfgs[slug] = build_config(slug, iphone_raw[slug], 'iphone', DEFAULT_SIM_IPHONE)

    # iPad
    ipad_raw = parse_ipad()
    ipad_order = ['ipad-11-2025', 'ipad-air-11-m3', 'ipad-air-11-m4', 'ipad-pro-11', 'ipad-pro-13']
    ipad_cfgs = OrderedDict()
    for slug in ipad_order:
        if slug in ipad_raw and ipad_raw[slug]:
            ipad_cfgs[slug] = build_config(slug, ipad_raw[slug], 'ipad', DEFAULT_SIM_WIFI)

    # MacBook
    mac_raw = parse_macbook()
    mac_order = ['mac-mini-m4', 'macbook-neo', 'macbook-air-13-m2', 'macbook-air-13-m5',
                 'macbook-air-15-m4', 'macbook-air-15-m5', 'macbook-pro-14-m5', 'macbook-pro-14-m5-pro']
    mac_cfgs = OrderedDict()
    for slug in mac_order:
        if slug in mac_raw and mac_raw[slug]:
            mac_cfgs[slug] = build_config(slug, mac_raw[slug], 'macbook', DEFAULT_SIM_NONE)

    # Watch
    watch_raw = parse_watch()
    watch_order = ['apple-watch-ultra-3', 'apple-watch-ultra-2', 'apple-watch-ultra-2025',
                   'apple-watch-series-11', 'apple-watch-series-10',
                   'apple-watch-se-3', 'apple-watch-se-2',
                   'samsung-galaxy-watch-8', 'samsung-galaxy-watch-classic-8']
    watch_cfgs = OrderedDict()
    for slug in watch_order:
        if slug in watch_raw and watch_raw[slug]:
            watch_cfgs[slug] = build_config(slug, watch_raw[slug], 'watch', DEFAULT_SIM_NONE)

    # AirPods
    airpods_raw = parse_airpods()
    airpods_order = ['airpods-max', 'airpods-pro-3', 'airpods-pro-2', 'airpods-4',
                     'galaxy-buds', 'marshall-headphones']
    airpods_cfgs = OrderedDict()
    for slug in airpods_order:
        if slug in airpods_raw and airpods_raw[slug]:
            airpods_cfgs[slug] = build_config(slug, airpods_raw[slug], 'airpods', DEFAULT_SIM_NONE)

    # Android
    android_raw = parse_android()
    android_order = ['samsung-galaxy-s26-ultra', 'samsung-galaxy-s26-plus', 'samsung-galaxy-s26',
                     'samsung-galaxy-s25-ultra', 'samsung-galaxy-s25-fe', 'samsung-galaxy-s25',
                     'samsung-galaxy-a56', 'samsung-galaxy-a36', 'samsung-galaxy-a26', 'samsung-galaxy-a17',
                     'samsung-galaxy-a07',
                     'xiaomi-mi-15t', 'xiaomi-redmi-note-15-pro', 'xiaomi-redmi-note-15',
                     'xiaomi-redmi-note-14s', 'xiaomi-redmi-note-14',
                     'meizu-note-21']
    android_cfgs = OrderedDict()
    for slug in android_order:
        if slug in android_raw and android_raw[slug]:
            android_cfgs[slug] = build_config(slug, android_raw[slug], 'android', DEFAULT_SIM_NONE)

    # Dyson
    dyson_raw = parse_dyson()
    dyson_order = ['dyson-hd18-pro', 'dyson-hd17', 'dyson-hd16',
                   'dyson-airwrap-hs09', 'dyson-airwrap-hs08', 'dyson-airwrap-hs05',
                   'dyson-airstrait-ht01',
                   'dyson-wash-g1', 'dyson-gen5-detect', 'dyson-v16s', 'dyson-v15s', 'dyson-v15',
                   'dyson-v12s', 'dyson-v12', 'dyson-v8']
    dyson_cfgs = OrderedDict()
    for slug in dyson_order:
        if slug in dyson_raw and dyson_raw[slug]:
            dyson_cfgs[slug] = build_config(slug, dyson_raw[slug], 'dyson', DEFAULT_SIM_NONE)

    # Audio
    audio_raw = parse_audio()
    audio_order = ['jbl-boombox-4', 'yandex-station-light-2', 'dji-mic-3', 'dji-mic-2', 'dji-mic-mini']
    audio_cfgs = OrderedDict()
    for slug in audio_order:
        if slug in audio_raw and audio_raw[slug]:
            audio_cfgs[slug] = build_config(slug, audio_raw[slug], 'audio', DEFAULT_SIM_NONE)

    # Записываем все
    Path(f'{OUT_DIR}/iphone-configs.ts').write_text(generate_category_ts('iphone', iphone_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/ipad-configs.ts').write_text(generate_category_ts('ipad', ipad_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/macbook-configs.ts').write_text(generate_category_ts('macbook', mac_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/watch-configs.ts').write_text(generate_category_ts('watch', watch_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/airpods-configs.ts').write_text(generate_category_ts('airpods', airpods_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/android-configs.ts').write_text(generate_category_ts('android', android_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/dyson-configs.ts').write_text(generate_category_ts('dyson', dyson_cfgs), encoding='utf-8')
    Path(f'{OUT_DIR}/audio-configs.ts').write_text(generate_category_ts('audio', audio_cfgs), encoding='utf-8')

    # Главный индекс
    index_lines = [
        '/** Автогенерация — индекс конфигов всех категорий */',
        'import type { ProductConfig } from "../product-configs";',
        'import { getIphoneConfig }  from "./iphone-configs";',
        'import { getIpadConfig }    from "./ipad-configs";',
        'import { getMacbookConfig } from "./macbook-configs";',
        'import { getWatchConfig }   from "./watch-configs";',
        'import { getAirpodsConfig } from "./airpods-configs";',
        'import { getAndroidConfig } from "./android-configs";',
        'import { getDysonConfig }   from "./dyson-configs";',
        'import { getAudioConfig }   from "./audio-configs";',
        '',
        'export { IPHONE_CONFIG_SLUGS }  from "./iphone-configs";',
        'export { IPAD_CONFIG_SLUGS }    from "./ipad-configs";',
        'export { MACBOOK_CONFIG_SLUGS } from "./macbook-configs";',
        'export { WATCH_CONFIG_SLUGS }   from "./watch-configs";',
        'export { AIRPODS_CONFIG_SLUGS } from "./airpods-configs";',
        'export { ANDROID_CONFIG_SLUGS } from "./android-configs";',
        'export { DYSON_CONFIG_SLUGS }   from "./dyson-configs";',
        'export { AUDIO_CONFIG_SLUGS }   from "./audio-configs";',
        '',
        'export function getProductConfig(category: string, slug: string): ProductConfig | undefined {',
        '  if (category === "iphone")  return getIphoneConfig(slug);',
        '  if (category === "ipad")    return getIpadConfig(slug);',
        '  if (category === "macbook") return getMacbookConfig(slug);',
        '  if (category === "watch")   return getWatchConfig(slug);',
        '  if (category === "airpods") return getAirpodsConfig(slug);',
        '  if (category === "android") return getAndroidConfig(slug);',
        '  if (category === "dyson")   return getDysonConfig(slug);',
        '  if (category === "audio")   return getAudioConfig(slug);',
        '  return undefined;',
        '}',
        '',
    ]
    Path(f'{OUT_DIR}/index.ts').write_text('\n'.join(index_lines), encoding='utf-8')

    print("\n=== Результаты ===\n")
    total = 0
    for cat_name, cfgs in [('iPhone', iphone_cfgs), ('iPad', ipad_cfgs), ('MacBook', mac_cfgs),
                            ('Watch', watch_cfgs), ('AirPods', airpods_cfgs), ('Android', android_cfgs),
                            ('Dyson', dyson_cfgs), ('Audio', audio_cfgs)]:
        print(f"📱 {cat_name}: {len(cfgs)} моделей")
        for slug, cfg in cfgs.items():
            nm = MODEL_NAMES.get(slug, slug)
            pf = f'{cfg["priceFrom"]:,}'.replace(',', ' ')
            print(f"   • {nm}: {len(cfg['colors'])}цв × {len(cfg['storage'])}конф × {len(cfg['sim'])}sim = {len(cfg['prices'])} | от {pf} ₽")
        total += len(cfgs)
        print()

    print(f"✅ ВСЕГО: {total} моделей в {OUT_DIR}/")


if __name__ == '__main__':
    main()
